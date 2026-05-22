import json
import sys
from datetime import datetime, time
from pathlib import Path


def fail(message, detail=None):
    print(json.dumps({"ok": False, "error": message, "detail": detail or ""}, ensure_ascii=False))
    sys.exit(1)


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value):
    return cell_text(value).lower()


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "date") and not isinstance(value, time):
        try:
            return value.date()
        except Exception:
            pass
    text = cell_text(value).replace("/", "-").replace("年", "-").replace("月", "-").replace("日", "")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_new_value(value):
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return ""
        try:
            number = float(text.replace(",", ""))
            return int(number) if number.is_integer() else number
        except ValueError:
            return value
    return value


def patch_mode(patch):
    mode = cell_text(patch.get("mode"))
    if mode:
        return mode
    if patch.get("oldValue") is not None and patch.get("newValue") is not None:
        return "text_replace"
    if patch.get("startDate") is not None and patch.get("endDate") is not None:
        return "date_range_set"
    return "text_replace"


def find_column(sheet, header_row, column_name):
    column = find_column_index(sheet, header_row, column_name)
    if column is not None:
        return column
    fail(f"找不到列: {column_name}", f"表头行: {header_row}")


def find_column_index(sheet, header_row, column_name):
    target = normalize_header(column_name)
    for column in range(1, sheet.max_column + 1):
        value = normalize_header(sheet.cell(row=header_row, column=column).value)
        if value == target:
            return column
    return None


def patch_date_range_set(workbook, patch):
    sheet_name = patch.get("sheetName") or workbook.sheetnames[0]
    if sheet_name not in workbook.sheetnames:
        fail(f"工作表不存在: {sheet_name}")
    sheet = workbook[sheet_name]
    header_row = int(patch.get("headerRowNumber") or 1)
    condition_column = find_column(sheet, header_row, patch.get("conditionColumn"))
    target_column = find_column(sheet, header_row, patch.get("targetColumn"))
    start_date = parse_date(patch.get("startDate"))
    end_date = parse_date(patch.get("endDate"))
    if not start_date or not end_date:
        fail("缺少合法日期范围")
    if end_date < start_date:
        fail("结束日期不能早于开始日期")
    new_value = parse_new_value(patch.get("newValue"))
    changed_cells = []
    for row in range(header_row + 1, sheet.max_row + 1):
        current_date = parse_date(sheet.cell(row=row, column=condition_column).value)
        if not current_date or current_date < start_date or current_date > end_date:
            continue
        cell = sheet.cell(row=row, column=target_column)
        old_value = cell.value
        cell.value = new_value
        changed_cells.append({
            "sheetName": sheet_name,
            "row": row,
            "column": target_column,
            "coordinate": cell.coordinate,
            "oldValue": cell_text(old_value),
            "newValue": cell_text(new_value),
        })
    if not changed_cells:
        fail("没有匹配到需要修改的单元格")
    return {
        "mode": "date_range_set",
        "sheetName": sheet_name,
        "headerRowNumber": header_row,
        "conditionColumn": patch.get("conditionColumn"),
        "targetColumn": patch.get("targetColumn"),
        "startDate": str(start_date),
        "endDate": str(end_date),
        "newValue": new_value,
        "changedCells": changed_cells[:50],
        "changedCellCount": len(changed_cells),
    }


def patch_text_replace(workbook, patch):
    old_value = cell_text(patch.get("oldValue"))
    if not old_value:
        fail("缺少 oldValue")
    if "newValue" not in patch:
        fail("缺少 newValue")
    new_value = patch.get("newValue")
    sheet_name = patch.get("sheetName") or ""
    target_sheet_names = [sheet_name] if sheet_name else list(workbook.sheetnames)
    target_column = cell_text(patch.get("targetColumn"))
    header_row = int(patch.get("headerRowNumber") or 0)
    if target_column and header_row <= 0:
        fail("指定 targetColumn 时必须同时提供 headerRowNumber")
    changed_cells = []
    for current_sheet_name in target_sheet_names:
        if current_sheet_name not in workbook.sheetnames:
            fail(f"工作表不存在: {current_sheet_name}")
        sheet = workbook[current_sheet_name]
        if target_column:
            column = find_column_index(sheet, header_row, target_column)
            if column is None:
                if sheet_name:
                    fail(f"找不到列: {target_column}", f"表头行: {header_row}")
                continue
            rows = range(header_row + 1, sheet.max_row + 1)
            cells = (sheet.cell(row=row, column=column) for row in rows)
        else:
            cells = (cell for row in sheet.iter_rows() for cell in row)
        for cell in cells:
            if cell_text(cell.value) != old_value:
                continue
            old_cell_value = cell.value
            cell.value = new_value
            changed_cells.append({
                "sheetName": current_sheet_name,
                "row": cell.row,
                "column": cell.column,
                "coordinate": cell.coordinate,
                "oldValue": cell_text(old_cell_value),
                "newValue": cell_text(new_value),
            })
    if not changed_cells:
        fail("没有匹配到需要修改的单元格")
    return {
        "mode": "text_replace",
        "sheetName": sheet_name or workbook.sheetnames[0],
        "targetColumn": target_column,
        "headerRowNumber": header_row or None,
        "oldValue": old_value,
        "newValue": new_value,
        "changedCells": changed_cells[:50],
        "changedCellCount": len(changed_cells),
    }


def patch_workbook(input_file, output_file, patch):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        fail(f"openpyxl 不可用，无法修改 Excel: {exc}")

    input_path = Path(input_file)
    output_path = Path(output_file)
    if input_path.suffix.lower() != ".xlsx":
        fail("格式保留型修改仅支持 .xlsx 文件")

    workbook = load_workbook(input_path, data_only=False)
    mode = patch_mode(patch)
    if mode == "text_replace":
        result = patch_text_replace(workbook, patch)
    elif mode == "date_range_set":
        result = patch_date_range_set(workbook, patch)
    else:
        fail(f"不支持的 patch mode: {mode}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return result


def main():
    if len(sys.argv) != 4:
        fail("usage: workbook_patch.py <input.xlsx> <output.xlsx> <json_patch>")
    try:
        patch = json.loads(sys.argv[3])
    except Exception as exc:
        fail(f"patch 不是合法 JSON: {exc}")
    result = patch_workbook(sys.argv[1], sys.argv[2], patch)
    print(json.dumps({"ok": True, "data": result}, ensure_ascii=False))


if __name__ == "__main__":
    main()
