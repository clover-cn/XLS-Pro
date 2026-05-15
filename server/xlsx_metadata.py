import json
import re
import sys
from pathlib import Path


def cell_to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def infer_type(values):
    filtered = [value for value in values if value not in (None, "")]
    if not filtered:
        return "empty"
    if all(isinstance(value, (int, float)) for value in filtered):
        return "number"
    if all(hasattr(value, "isoformat") for value in filtered):
        return "date"
    return "text"


HEADER_KEYWORDS = [
    "日期", "时间", "摘要", "科目", "借方", "贷方", "金额", "用量", "数量", "次数",
    "凭证", "编码", "名称", "客户", "供应商", "单位", "收入", "支出", "余额",
]


def cell_kind(value):
    if value is None or value == "":
        return "empty"
    if isinstance(value, (int, float)):
        return "number"
    if hasattr(value, "isoformat"):
        return "date"
    text = cell_to_text(value)
    if re.fullmatch(r"-?\d+(?:,\d{3})*(?:\.\d+)?", text):
        return "number"
    if re.fullmatch(r"\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?", text):
        return "date"
    return "text"


def header_data_fit(header_value, data_value):
    header_text = cell_to_text(header_value)
    if not header_text:
        return 0
    data_kind = cell_kind(data_value)
    if data_kind in ("date", "number"):
        return 4
    if data_kind == "text" and header_text != cell_to_text(data_value):
        return 1
    return 0


def best_header_index(raw_rows):
    best_index = 0
    best_score = -1
    for index, row in enumerate(raw_rows):
        values = [cell_to_text(value) for value in row]
        non_empty = [value for value in values if value]
        next_rows = raw_rows[index + 1: index + 4]
        keyword_score = sum(3 for value in non_empty if any(keyword in value for keyword in HEADER_KEYWORDS))
        data_score = 0
        for next_row in next_rows:
            for column_index, value in enumerate(values):
                data_score += header_data_fit(value, next_row[column_index] if column_index < len(next_row) else "")
        score = len(non_empty) + keyword_score + data_score
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def main():
    if len(sys.argv) not in (2, 3):
        raise SystemExit("usage: xlsx_metadata.py <file> [preview_rows]")

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise SystemExit(f"openpyxl 不可用，无法解析 .xlsx 元数据: {exc}")

    file_path = Path(sys.argv[1])
    preview_rows = max(1, min(int(sys.argv[2]) if len(sys.argv) == 3 else 3, 50))
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active

    max_column = sheet.max_column or 0
    raw_rows = []
    for row_number in range(1, min(sheet.max_row or 0, preview_rows) + 1):
        values = [cell_to_text(sheet.cell(row=row_number, column=column).value) for column in range(1, max_column + 1)]
        raw_rows.append({"rowNumber": row_number, "values": values})

    merged_cells = []
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= preview_rows:
            merged_cells.append({
                "range": str(merged_range),
                "value": cell_to_text(sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value),
            })

    raw_values = [row["values"] for row in raw_rows]
    header_index = best_header_index(raw_values) if raw_values else 0
    headers = raw_values[header_index] if raw_values else []
    scan_values = [[] for _ in headers]

    for row in sheet.iter_rows(min_row=header_index + 2, max_row=min(sheet.max_row or 0, header_index + 51), values_only=True):
        for index, value in enumerate(list(row)[: len(headers)]):
            scan_values[index].append(value)

    payload = {
        "fileKind": "xlsx",
        "sheetName": sheet.title,
        "sheetNames": workbook.sheetnames,
        "totalRows": sheet.max_row or 0,
        "totalColumns": max_column,
        "previewRows": preview_rows,
        "rawRows": raw_rows,
        "mergedCells": merged_cells,
        "detectedHeaderRowNumber": header_index + 1 if raw_rows else None,
        "columns": [
            {"name": header or f"Column {index + 1}", "type": infer_type(scan_values[index])}
            for index, header in enumerate(headers)
        ],
    }
    print(json.dumps(payload, ensure_ascii=False))


if __name__ == "__main__":
    main()
