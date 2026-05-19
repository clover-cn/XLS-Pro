import csv
import json
import random
import re
import sys
from pathlib import Path


MAX_READ_ROWS = 200
MAX_SEARCH_RESULTS = 50
MAX_FILTER_RESULTS = 200
MAX_CELL_TEXT = 200
BATCH_SIZE = 1000
PROGRESS_EVERY_ROWS = 1000
WORKBOOK_INDEX_VERSION = 2
HEADER_KEYWORDS = [
    "日期", "时间", "摘要", "科目", "借方", "贷方", "金额", "用量", "数量", "次数",
    "凭证", "编码", "名称", "客户", "供应商", "单位", "收入", "支出", "余额",
]


def cell_to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def trim_text(value, limit=MAX_CELL_TEXT):
    text = cell_to_text(value)
    return text if len(text) <= limit else f"{text[:limit]}..."


def fail(message):
    print(json.dumps({"ok": False, "error": message}, ensure_ascii=False))
    sys.exit(1)


def emit_event(event, **payload):
    print(json.dumps({"event": event, **payload}, ensure_ascii=False), flush=True)


def load_duckdb():
    try:
        import duckdb
    except Exception as exc:
        fail(f"duckdb 不可用，请先安装: pip install duckdb。原始错误: {exc}")
    return duckdb


def load_openpyxl():
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        fail(f"openpyxl 不可用，请先安装: pip install openpyxl。原始错误: {exc}")
    return load_workbook


def quote_ident(value):
    return '"' + str(value).replace('"', '""') + '"'


def normalize_sheet_name(value, fallback):
    text = cell_to_text(value)
    return text or fallback


def infer_type(values):
    filtered = [value for value in values if value not in ("", None)]
    if not filtered:
        return "empty"
    numeric = 0
    for value in filtered:
        try:
            float(str(value).replace(",", ""))
            numeric += 1
        except Exception:
            pass
    if numeric >= max(1, int(len(filtered) * 0.8)):
        return "number"
    return "text"


def cell_kind(value):
    text = cell_to_text(value)
    if not text:
        return "empty"
    if re.fullmatch(r"-?\d+(?:,\d{3})*(?:\.\d+)?", text):
        return "number"
    if re.fullmatch(r"\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?", text):
        return "date"
    return "text"


def header_data_fit(header_value, data_value):
    header_text = cell_to_text(header_value)
    if not header_text:
        return 0
    data_kind = cell_kind(data_value)
    if data_kind in ("date", "number"):
        return 4
    if data_kind == "text" and header_text != cell_to_text(data_value):
        return 1
    return 0


def best_header_row(raw_rows):
    if not raw_rows:
        return None
    best = (0, -1)
    for index, row in enumerate(raw_rows):
        values = [cell_to_text(value) for value in row]
        non_empty = [value for value in values if value]
        next_rows = raw_rows[index + 1: index + 4]
        keyword_score = sum(3 for value in non_empty if any(keyword in value for keyword in HEADER_KEYWORDS))
        data_score = 0
        for next_row in next_rows:
            for column_index, value in enumerate(values):
                data_score += header_data_fit(value, next_row[column_index] if column_index < len(next_row) else "")
        score = len(non_empty) + keyword_score + data_score
        if score > best[1]:
            best = (index, score)
    if best[1] <= 0:
        return None
    return best[0] + 1


def make_columns(total_columns, raw_rows):
    header_row = best_header_row(raw_rows)
    header_values = raw_rows[header_row - 1] if header_row and header_row <= len(raw_rows) else []
    sample_start = header_row if header_row is not None else 0
    columns = []
    seen = {}
    for index in range(total_columns):
        name = cell_to_text(header_values[index] if index < len(header_values) else "") or f"Column {index + 1}"
        count = seen.get(name, 0) + 1
        seen[name] = count
        if count > 1:
            name = f"{name}_{count}"
        sample_values = [
            row[index] if index < len(row) else ""
            for row in raw_rows[sample_start: min(len(raw_rows), sample_start + 50)]
        ]
        columns.append({
            "index": index + 1,
            "storageName": f"c{index + 1}",
            "name": name,
            "type": infer_type(sample_values),
        })
    return header_row, columns


def create_sheet_table(connection, table_name, total_columns):
    columns_sql = ", ".join([f"c{index} VARCHAR" for index in range(1, total_columns + 1)])
    connection.execute(f"DROP TABLE IF EXISTS {quote_ident(table_name)}")
    connection.execute(f"CREATE TABLE {quote_ident(table_name)} (row_number BIGINT, {columns_sql})")


def insert_rows(connection, table_name, rows, total_columns):
    if not rows:
        return
    placeholders = ", ".join(["?"] * (total_columns + 1))
    sql = f"INSERT INTO {quote_ident(table_name)} VALUES ({placeholders})"
    connection.executemany(sql, rows)


def iter_csv_rows(file_path):
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.reader(handle):
            yield [cell_to_text(value) for value in row]


def iter_xlsx_sheets(file_path):
    load_workbook = load_openpyxl()
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        yield sheet.title, sheet.max_column or 1, sheet.max_row or None, ([cell_to_text(value) for value in row] for row in sheet.iter_rows(values_only=True))


def csv_total_columns(file_path):
    total_columns = 0
    for row in iter_csv_rows(file_path):
        total_columns = max(total_columns, len(row))
    return total_columns or 1


def index_sheet(connection, table_name, row_iterable, total_columns_hint=None, sheet_name="", total_rows_hint=None):
    buffered_raw = []
    total_columns = total_columns_hint or 0
    rows_seen = 0
    batch = []
    if total_columns_hint:
        create_sheet_table(connection, table_name, total_columns_hint)
    emit_event("index_progress", sheetName=sheet_name, indexedRows=0, totalRows=total_rows_hint, phase="started")
    for row in row_iterable:
        rows_seen += 1
        total_columns = max(total_columns, len(row))
        if len(buffered_raw) < 80:
            buffered_raw.append(row)
        batch.append(row)
        if len(batch) >= BATCH_SIZE:
            if rows_seen == len(batch) and not total_columns_hint:
                create_sheet_table(connection, table_name, total_columns or 1)
            normalized = []
            for offset, batch_row in enumerate(batch, start=rows_seen - len(batch) + 1):
                normalized.append([offset] + [cell_to_text(batch_row[index]) if index < len(batch_row) else "" for index in range(total_columns or 1)])
            insert_rows(connection, table_name, normalized, total_columns or 1)
            batch = []
            if rows_seen % PROGRESS_EVERY_ROWS == 0:
                emit_event("index_progress", sheetName=sheet_name, indexedRows=rows_seen, totalRows=total_rows_hint, phase="indexing")

    if rows_seen == 0:
        total_columns = 1
        create_sheet_table(connection, table_name, total_columns)
    elif rows_seen <= len(batch) and not total_columns_hint:
        create_sheet_table(connection, table_name, total_columns or 1)

    if batch:
        start = rows_seen - len(batch) + 1
        normalized = []
        for offset, batch_row in enumerate(batch, start=start):
            normalized.append([offset] + [cell_to_text(batch_row[index]) if index < len(batch_row) else "" for index in range(total_columns or 1)])
        insert_rows(connection, table_name, normalized, total_columns or 1)
    emit_event("index_progress", sheetName=sheet_name, indexedRows=rows_seen, totalRows=total_rows_hint, phase="completed")

    header_row, columns = make_columns(total_columns or 1, buffered_raw)
    return {
        "totalRows": rows_seen,
        "totalColumns": total_columns or 1,
        "detectedHeaderRowNumber": header_row,
        "columns": columns,
        "rawRows": [
            {"rowNumber": index + 1, "values": [trim_text(value) for value in row[: min(total_columns or 1, 50)]]}
            for index, row in enumerate(buffered_raw[:20])
        ],
    }


def build_index(file_path, index_dir):
    duckdb = load_duckdb()
    index_dir.mkdir(parents=True, exist_ok=True)
    db_path = index_dir / "workbook.duckdb"
    manifest_path = index_dir / "manifest.json"
    if db_path.exists():
        db_path.unlink()
    connection = duckdb.connect(str(db_path))
    sheets = []
    ext = file_path.suffix.lower()
    if ext == ".csv":
        table_name = "sheet_0"
        profile = index_sheet(connection, table_name, iter_csv_rows(file_path), csv_total_columns(file_path), "CSV")
        sheets.append({"sheetName": "CSV", "tableName": table_name, **profile})
    elif ext == ".xlsx":
        for index, (sheet_name, total_columns, total_rows, rows) in enumerate(iter_xlsx_sheets(file_path)):
            table_name = f"sheet_{index}"
            profile = index_sheet(connection, table_name, rows, total_columns, normalize_sheet_name(sheet_name, f"Sheet{index + 1}"), total_rows)
            sheets.append({"sheetName": normalize_sheet_name(sheet_name, f"Sheet{index + 1}"), "tableName": table_name, **profile})
    else:
        fail("仅支持 .csv 和 .xlsx 文件")
    connection.close()
    manifest = {
        "version": WORKBOOK_INDEX_VERSION,
        "sourceFile": file_path.name,
        "dbFile": db_path.name,
        "sheetNames": [sheet["sheetName"] for sheet in sheets],
        "sheets": sheets,
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def load_manifest(index_dir):
    manifest_path = index_dir / "manifest.json"
    if not manifest_path.is_file():
        fail("索引不存在，请先构建 workbook 索引")
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def connect_index(index_dir):
    duckdb = load_duckdb()
    return duckdb.connect(str(index_dir / "workbook.duckdb"), read_only=True)


def find_sheet(manifest, sheet_name=None):
    sheets = manifest.get("sheets", [])
    if not sheets:
        fail("索引中没有工作表")
    if not sheet_name:
        return sheets[0]
    for sheet in sheets:
        if sheet["sheetName"] == sheet_name:
            return sheet
    fail(f"工作表不存在: {sheet_name}")


def column_storage(sheet, column):
    if column is None or column == "":
        fail("缺少列名或列号")
    text = str(column).strip()
    if re.fullmatch(r"c\d+", text):
        index = int(text[1:])
        if 1 <= index <= sheet["totalColumns"]:
            return text
    if text.isdigit():
        index = int(text)
        if 1 <= index <= sheet["totalColumns"]:
            return f"c{index}"
    for item in sheet.get("columns", []):
        if item["name"] == text or item["storageName"] == text:
            return item["storageName"]
    fail(f"列不存在: {column}")


def row_to_payload(row_number, values):
    return {"rowNumber": int(row_number), "values": [trim_text(value) for value in values]}


def read_rows(connection, sheet, start_row, end_row):
    start_row = int(start_row)
    end_row = int(end_row)
    if start_row < 1 or end_row < start_row:
        fail("行号必须是 1-based，且 endRow 不能小于 startRow")
    if end_row - start_row + 1 > MAX_READ_ROWS:
        fail(f"单次最多读取 {MAX_READ_ROWS} 行")
    storage_columns = [f"c{index}" for index in range(1, sheet["totalColumns"] + 1)]
    sql = f"SELECT row_number, {', '.join(map(quote_ident, storage_columns))} FROM {quote_ident(sheet['tableName'])} WHERE row_number BETWEEN ? AND ? ORDER BY row_number"
    rows = connection.execute(sql, [start_row, end_row]).fetchall()
    return {
        "sheetName": sheet["sheetName"],
        "startRow": start_row,
        "endRow": end_row,
        "rows": [row_to_payload(row[0], row[1:]) for row in rows],
    }


def sample_rows(connection, sheet, arguments):
    mode = arguments.get("mode", "first")
    count = max(1, min(int(arguments.get("count", 20)), MAX_READ_ROWS))
    if mode == "last":
        end_row = sheet["totalRows"]
        start_row = max(1, end_row - count + 1)
    elif mode == "around":
        center = int(arguments.get("rowNumber", 1))
        start_row = max(1, center - count // 2)
        end_row = min(sheet["totalRows"], start_row + count - 1)
    elif mode == "random":
        max_start = max(1, sheet["totalRows"] - count + 1)
        start_row = random.randint(1, max_start)
        end_row = min(sheet["totalRows"], start_row + count - 1)
    else:
        start_row = 1
        end_row = min(sheet["totalRows"], count)
    return read_rows(connection, sheet, start_row, end_row)


def search(connection, manifest, arguments):
    query = cell_to_text(arguments.get("query"))
    if not query:
        fail("搜索关键词不能为空")
    max_results = max(1, min(int(arguments.get("maxResults", 20)), MAX_SEARCH_RESULTS))
    sheets = [find_sheet(manifest, arguments.get("sheetName"))] if arguments.get("sheetName") else manifest.get("sheets", [])
    needle = f"%{query.casefold()}%"
    results = []
    for sheet in sheets:
        columns = [f"c{index}" for index in range(1, sheet["totalColumns"] + 1)]
        where = " OR ".join([f"lower({quote_ident(column)}) LIKE ?" for column in columns])
        sql = f"SELECT row_number, {', '.join(map(quote_ident, columns))} FROM {quote_ident(sheet['tableName'])} WHERE {where} ORDER BY row_number LIMIT ?"
        rows = connection.execute(sql, [needle] * len(columns) + [max_results]).fetchall()
        for row in rows:
            row_number = int(row[0])
            for index, value in enumerate(row[1:], start=1):
                text = cell_to_text(value)
                if query.casefold() in text.casefold():
                    results.append({
                        "sheetName": sheet["sheetName"],
                        "rowNumber": row_number,
                        "columnNumber": index,
                        "columnName": sheet["columns"][index - 1]["name"] if index - 1 < len(sheet["columns"]) else f"Column {index}",
                        "value": trim_text(text),
                    })
                    if len(results) >= max_results:
                        return {"query": query, "resultCount": len(results), "results": results}
    return {"query": query, "resultCount": len(results), "results": results}


def filter_rows(connection, sheet, arguments):
    column = column_storage(sheet, arguments.get("column"))
    operator = arguments.get("operator", "contains")
    value = cell_to_text(arguments.get("value"))
    max_results = max(1, min(int(arguments.get("maxResults", 50)), MAX_FILTER_RESULTS))
    params = []
    if operator == "equals":
        predicate = f"{quote_ident(column)} = ?"
        params.append(value)
    elif operator == "not_empty":
        predicate = f"{quote_ident(column)} <> ''"
    elif operator in ("gt", "gte", "lt", "lte"):
        op = {"gt": ">", "gte": ">=", "lt": "<", "lte": "<="}[operator]
        predicate = f"try_cast(replace({quote_ident(column)}, ',', '') AS DOUBLE) {op} ?"
        params.append(float(value))
    else:
        predicate = f"lower({quote_ident(column)}) LIKE ?"
        params.append(f"%{value.casefold()}%")
    columns = [f"c{index}" for index in range(1, sheet["totalColumns"] + 1)]
    sql = f"SELECT row_number, {', '.join(map(quote_ident, columns))} FROM {quote_ident(sheet['tableName'])} WHERE {predicate} ORDER BY row_number LIMIT ?"
    rows = connection.execute(sql, params + [max_results]).fetchall()
    return {
        "sheetName": sheet["sheetName"],
        "matchedRows": len(rows),
        "rows": [row_to_payload(row[0], row[1:]) for row in rows],
    }


def aggregate(connection, sheet, arguments):
    group_by = arguments.get("groupBy")
    metric_column = column_storage(sheet, arguments.get("column"))
    operation = arguments.get("operation", "sum")
    if operation not in ("sum", "avg", "min", "max", "count"):
        fail("聚合 operation 仅支持 sum/avg/min/max/count")
    value_expr = f"try_cast(replace({quote_ident(metric_column)}, ',', '') AS DOUBLE)"
    if operation == "count":
        metric_expr = "count(*)"
    else:
        metric_expr = f"{operation}({value_expr})"
    if group_by:
        group_column = column_storage(sheet, group_by)
        sql = f"SELECT {quote_ident(group_column)} AS group_value, {metric_expr} AS metric_value FROM {quote_ident(sheet['tableName'])} GROUP BY {quote_ident(group_column)} ORDER BY metric_value DESC NULLS LAST LIMIT 100"
    else:
        sql = f"SELECT {metric_expr} AS metric_value FROM {quote_ident(sheet['tableName'])}"
    rows = connection.execute(sql).fetchall()
    return {
        "sheetName": sheet["sheetName"],
        "operation": operation,
        "column": arguments.get("column"),
        "groupBy": group_by or "",
        "rows": [{"group": row[0], "value": row[1]} for row in rows] if group_by else [{"value": rows[0][0] if rows else None}],
    }


def profile_column(connection, sheet, arguments):
    column = column_storage(sheet, arguments.get("column"))
    qcol = quote_ident(column)
    total, non_empty, distinct_count, numeric_count, min_value, max_value = connection.execute(
        f"SELECT count(*), count(NULLIF({qcol}, '')), count(DISTINCT NULLIF({qcol}, '')), count(try_cast(replace({qcol}, ',', '') AS DOUBLE)), min(try_cast(replace({qcol}, ',', '') AS DOUBLE)), max(try_cast(replace({qcol}, ',', '') AS DOUBLE)) FROM {quote_ident(sheet['tableName'])}"
    ).fetchone()
    samples = connection.execute(
        f"SELECT {qcol}, count(*) AS n FROM {quote_ident(sheet['tableName'])} WHERE {qcol} <> '' GROUP BY {qcol} ORDER BY n DESC LIMIT 20"
    ).fetchall()
    return {
        "sheetName": sheet["sheetName"],
        "column": arguments.get("column"),
        "totalRows": total,
        "nonEmptyRows": non_empty,
        "distinctCount": distinct_count,
        "numericCount": numeric_count,
        "numericMin": min_value,
        "numericMax": max_value,
        "topValues": [{"value": trim_text(row[0]), "count": row[1]} for row in samples],
    }


def run_tool(index_dir, tool_name, arguments):
    manifest = load_manifest(index_dir)
    if tool_name in ("excel_list_sheets", "excel_describe_workbook"):
        return {
            "sheetNames": manifest.get("sheetNames", []),
            "sheets": [{
                "sheetName": sheet["sheetName"],
                "totalRows": sheet["totalRows"],
                "totalColumns": sheet["totalColumns"],
                "detectedHeaderRowNumber": sheet.get("detectedHeaderRowNumber"),
            } for sheet in manifest.get("sheets", [])],
        }
    sheet = find_sheet(manifest, arguments.get("sheetName"))
    if tool_name in ("excel_get_schema",):
        return {"sheetName": sheet["sheetName"], "totalRows": sheet["totalRows"], "totalColumns": sheet["totalColumns"], "detectedHeaderRowNumber": sheet.get("detectedHeaderRowNumber"), "columns": sheet.get("columns", []), "rawRows": sheet.get("rawRows", [])}
    connection = connect_index(index_dir)
    try:
        if tool_name in ("excel_read_rows",):
            data = read_rows(connection, sheet, arguments.get("startRow", 1), arguments.get("endRow", 1))
        elif tool_name == "excel_sample_rows":
            data = sample_rows(connection, sheet, arguments)
        elif tool_name == "excel_search":
            data = search(connection, manifest, arguments)
        elif tool_name == "excel_filter_rows":
            data = filter_rows(connection, sheet, arguments)
        elif tool_name == "excel_aggregate":
            data = aggregate(connection, sheet, arguments)
        elif tool_name == "excel_profile_column":
            data = profile_column(connection, sheet, arguments)
        else:
            fail(f"未知 Excel 工具: {tool_name}")
    finally:
        connection.close()
    return data


def main():
    if len(sys.argv) < 2:
        fail("usage: excel_tools.py build-index|tool ...")

    mode = sys.argv[1]
    try:
        if mode == "build-index":
            if len(sys.argv) != 4:
                fail("usage: excel_tools.py build-index <file> <index_dir>")
            data = build_index(Path(sys.argv[2]).resolve(), Path(sys.argv[3]).resolve())
        elif mode == "tool":
            if len(sys.argv) != 5:
                fail("usage: excel_tools.py tool <index_dir> <tool_name> <json_arguments>")
            arguments = json.loads(sys.argv[4] or "{}")
            data = run_tool(Path(sys.argv[2]).resolve(), sys.argv[3], arguments)
        else:
            if len(sys.argv) != 4:
                fail("legacy usage: excel_tools.py <file> <tool_name> <json_arguments>")
            file_path = Path(sys.argv[1]).resolve()
            index_dir = file_path.parent / ".excel-index-smoke"
            build_index(file_path, index_dir)
            data = run_tool(index_dir, sys.argv[2], json.loads(sys.argv[3] or "{}"))
    except json.JSONDecodeError as exc:
        fail(f"工具参数不是合法 JSON: {exc}")
    except SystemExit:
        raise
    except Exception as exc:
        fail(str(exc))

    print(json.dumps({"ok": True, "data": data}, ensure_ascii=False))


if __name__ == "__main__":
    main()
